from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database.audit import recent_audit_logs
from database.context import default_establishment_id, is_all_establishments_scope
from database.db import EXPORTS_DIR, get_connection
from database.queries import CORRECTED, REFUSED, VALIDATED
from utils.anomaly_detection import detect_anomalies
from utils.app_logging import log_info
from utils.time_utils import week_bounds

EXPORT_DIR = EXPORTS_DIR

SUMMARY_COLUMNS = [
    "Employé",
    "Objectif semaine",
    "Total heures validées",
    "Écart",
    "Statut",
    "Heures restaurant",
    "Heures ménage",
    "Heures entretien",
    "Nombre de créneaux",
    "Nombre de demandes manuelles acceptées",
    "Nombre d'anomalies",
    "Commentaire responsable",
]

DETAIL_COLUMNS = [
    "Date",
    "Employé",
    "Service",
    "Début",
    "Fin",
    "Durée",
    "Origine",
    "Statut",
    "Commentaire responsable",
]

MANUAL_COLUMNS = [
    "Date",
    "Employé",
    "Service",
    "Début demandé",
    "Fin demandée",
    "Durée demandée",
    "Raison",
    "Statut",
    "Commentaire responsable",
]

ANOMALY_COLUMNS = ["Date", "Employé", "Service", "Type anomalie", "Détail", "Statut"]

LOG_COLUMNS = [
    "Date action",
    "Employé",
    "Service",
    "Action",
    "Ancienne valeur",
    "Nouvelle valeur",
    "Commentaire responsable",
]

ACTIVITY_LOG_COLUMNS = [
    "Date action",
    "Responsable",
    "Établissement",
    "Action",
    "Élément",
    "Ancienne valeur",
    "Nouvelle valeur",
    "Commentaire",
]

WEEK_DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
WEEK_VIEW_COLUMNS = [
    "Employé",
    *WEEK_DAYS,
    "Total semaine",
    "Objectif semaine",
    "Heures validées",
    "Écart",
    "Statut",
    "Services travaillés",
]
WEEK_DETAIL_COLUMNS = ["Employé", *WEEK_DAYS]


def export_weekly_excel(establishment_id: int | None = None) -> Path:
    establishment_id = establishment_id or default_establishment_id()
    EXPORT_DIR.mkdir(exist_ok=True)
    week_start, week_end = week_bounds()
    output = EXPORT_DIR / f"campflow_export_semaine_{week_start[:10]}.xlsx"

    sessions = _weekly_sessions(week_start, week_end, establishment_id)
    manual_requests = _weekly_manual_requests(week_start, week_end, establishment_id)
    logs = _weekly_validation_logs(week_start, week_end, establishment_id)
    activity_logs = recent_audit_logs(establishment_id, limit=200)
    anomalies = detect_anomalies(sessions, manual_requests)
    employees = _active_employees(establishment_id)

    sheets = {
        "Vue semaine": _build_week_view(sessions, week_start, employees),
        "Vue semaine détaillée": _build_week_detail_view(sessions, week_start, employees),
        "Synthèse semaine": _build_summary(sessions, manual_requests, anomalies, employees),
        "Détail créneaux": _build_details(sessions),
        "Demandes manuelles": _build_manual_requests(manual_requests),
        "Anomalies": _build_anomalies(anomalies),
        "Journal validation": _build_logs(logs),
        "Journal activité": _build_activity_logs(activity_logs),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_workbook(writer.book)

    log_info(f"Export généré : Excel {output.name}")
    return output


def export_weekly_csv(establishment_id: int | None = None) -> list[Path]:
    establishment_id = establishment_id or default_establishment_id()
    EXPORT_DIR.mkdir(exist_ok=True)
    week_start, week_end = week_bounds()
    suffix = week_start[:10]

    sessions = _weekly_sessions(week_start, week_end, establishment_id)
    manual_requests = _weekly_manual_requests(week_start, week_end, establishment_id)
    anomalies = detect_anomalies(sessions, manual_requests)
    employees = _active_employees(establishment_id)

    exports = {
        f"campflow_synthese_semaine_{suffix}.csv": _build_summary_csv(sessions, manual_requests, anomalies, employees),
        f"campflow_detail_creneaux_{suffix}.csv": _build_details(sessions),
        f"campflow_demandes_manuelles_{suffix}.csv": _build_manual_requests(manual_requests),
    }

    paths = []
    for filename, df in exports.items():
        path = EXPORT_DIR / filename
        df.to_csv(path, index=False, sep=";", encoding="utf-8")
        paths.append(path)
    log_info("Export généré : CSV " + ", ".join(path.name for path in paths))
    return paths


def _weekly_sessions(week_start: str, week_end: str, establishment_id: int) -> pd.DataFrame:
    scope = "" if is_all_establishments_scope(establishment_id) else "ws.establishment_id = ? AND"
    params = (week_start, week_end) if is_all_establishments_scope(establishment_id) else (establishment_id, week_start, week_end)
    with get_connection() as conn:
        return pd.read_sql_query(
            f"""
            SELECT
                ws.id,
                ws.employee_id,
                e.first_name || ' ' || e.last_name AS employee,
                s.name AS service,
                ws.start_time,
                ws.end_time,
                ws.duration_hours,
                ws.corrected_duration_hours,
                ws.source,
                ws.validation_status,
                ws.manager_comment
            FROM work_sessions ws
            JOIN employees e ON e.id = ws.employee_id
            JOIN services s ON s.id = ws.service_id
            WHERE {scope} ws.start_time >= ? AND ws.start_time < ?
            ORDER BY ws.start_time ASC
            """,
            conn,
            params=params,
        )


def _weekly_manual_requests(week_start: str, week_end: str, establishment_id: int) -> pd.DataFrame:
    scope = "" if is_all_establishments_scope(establishment_id) else "m.establishment_id = ? AND"
    params = (week_start, week_end) if is_all_establishments_scope(establishment_id) else (establishment_id, week_start, week_end)
    with get_connection() as conn:
        return pd.read_sql_query(
            f"""
            SELECT
                m.id,
                m.employee_id,
                e.first_name || ' ' || e.last_name AS employee,
                s.name AS service,
                m.requested_start_time,
                m.requested_end_time,
                m.requested_duration_hours,
                m.reason,
                m.status,
                m.manager_comment,
                m.created_at,
                m.reviewed_at
            FROM manual_time_requests m
            JOIN employees e ON e.id = m.employee_id
            JOIN services s ON s.id = m.service_id
            WHERE {scope} m.created_at >= ? AND m.created_at < ?
            ORDER BY m.created_at ASC
            """,
            conn,
            params=params,
        )


def _weekly_validation_logs(week_start: str, week_end: str, establishment_id: int) -> pd.DataFrame:
    scope = "" if is_all_establishments_scope(establishment_id) else "vl.establishment_id = ? AND"
    params = (week_start, week_end) if is_all_establishments_scope(establishment_id) else (establishment_id, week_start, week_end)
    with get_connection() as conn:
        return pd.read_sql_query(
            f"""
            SELECT
                vl.timestamp,
                COALESCE(ews.first_name || ' ' || ews.last_name, em.first_name || ' ' || em.last_name) AS employee,
                COALESCE(sws.name, sm.name) AS service,
                vl.action,
                vl.old_value,
                vl.new_value,
                vl.manager_comment
            FROM validation_logs vl
            LEFT JOIN work_sessions ws ON ws.id = vl.work_session_id
            LEFT JOIN employees ews ON ews.id = ws.employee_id
            LEFT JOIN services sws ON sws.id = ws.service_id
            LEFT JOIN manual_time_requests m ON m.id = vl.manual_request_id
            LEFT JOIN employees em ON em.id = m.employee_id
            LEFT JOIN services sm ON sm.id = m.service_id
            WHERE {scope} vl.timestamp >= ? AND vl.timestamp < ?
            ORDER BY vl.timestamp ASC
            """,
            conn,
            params=params,
        )


def _active_employees(establishment_id: int) -> pd.DataFrame:
    scope = "" if is_all_establishments_scope(establishment_id) else "AND establishment_id = ?"
    params = () if is_all_establishments_scope(establishment_id) else (establishment_id,)
    with get_connection() as conn:
        return pd.read_sql_query(
            f"""
            SELECT
                id AS employee_id,
                first_name || ' ' || last_name AS employee,
                COALESCE(weekly_target_hours, 35) AS weekly_target_hours
            FROM employees
            WHERE active = 1 {scope}
            ORDER BY last_name, first_name
            """,
            conn,
            params=params,
        )


def _build_week_view(sessions: pd.DataFrame, week_start: str, employees: pd.DataFrame | None = None) -> pd.DataFrame:
    employees = _employees_for_export(sessions, employees)
    if employees.empty:
        return pd.DataFrame(columns=WEEK_VIEW_COLUMNS)

    rows = []
    payable = _payable_sessions(sessions)
    day_by_date = _week_day_by_date(week_start)

    if not payable.empty:
        payable.loc[:, "date"] = payable["start_time"].map(_date)
        payable.loc[:, "day"] = payable["date"].map(day_by_date)
        payable = payable.dropna(subset=["day"])
        totals = payable.groupby(["employee_id", "day"], as_index=False)["paid_hours"].sum()
        services = payable.groupby("employee_id")["service"].apply(
            lambda values: ", ".join(sorted({str(value) for value in values if str(value).strip()}))
        )
    else:
        totals = pd.DataFrame(columns=["employee_id", "day", "paid_hours"])
        services = pd.Series(dtype="object")

    for employee in employees.itertuples(index=False):
        employee_totals = totals[totals["employee_id"] == employee.employee_id]
        by_day = dict(zip(employee_totals["day"], employee_totals["paid_hours"], strict=False))
        target = _number_value(getattr(employee, "weekly_target_hours", 35), 35)
        row = {"Employé": employee.employee}
        for day in WEEK_DAYS:
            row[day] = round(float(by_day.get(day, 0)), 2)
        row["Total semaine"] = round(sum(row[day] for day in WEEK_DAYS), 2)
        row["Objectif semaine"] = round(target, 2)
        row["Heures validées"] = row["Total semaine"]
        row["Écart"] = round(row["Heures validées"] - target, 2)
        row["Statut"] = _hours_status(row["Écart"])
        row["Services travaillés"] = services.get(employee.employee_id, "") if not services.empty else ""
        rows.append(row)

    return pd.DataFrame(rows, columns=WEEK_VIEW_COLUMNS)


def _build_week_detail_view(sessions: pd.DataFrame, week_start: str, employees: pd.DataFrame | None = None) -> pd.DataFrame:
    employees = _employees_for_export(sessions, employees)
    if employees.empty:
        return pd.DataFrame(columns=WEEK_DETAIL_COLUMNS)

    payable = _payable_sessions(sessions)
    day_by_date = _week_day_by_date(week_start)
    rows = []

    if not payable.empty:
        payable.loc[:, "date"] = payable["start_time"].map(_date)
        payable.loc[:, "day"] = payable["date"].map(day_by_date)
        payable = payable.dropna(subset=["day"]).sort_values(["employee", "start_time"])

    for employee in employees.itertuples(index=False):
        employee_sessions = payable[payable["employee_id"] == employee.employee_id] if not payable.empty else payable
        row = {"Employé": employee.employee}
        for day in WEEK_DAYS:
            day_sessions = employee_sessions[employee_sessions["day"] == day] if not employee_sessions.empty else employee_sessions
            row[day] = "\n".join(_session_detail(item) for item in day_sessions.itertuples(index=False))
        rows.append(row)

    return pd.DataFrame(rows, columns=WEEK_DETAIL_COLUMNS)


def _build_summary(
    sessions: pd.DataFrame,
    manual_requests: pd.DataFrame,
    anomalies: pd.DataFrame,
    employees: pd.DataFrame | None = None,
) -> pd.DataFrame:
    employees = _employees_for_export(sessions, employees)
    if employees.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    validated = sessions[sessions["validation_status"].isin([VALIDATED, CORRECTED])].copy()

    if validated.empty:
        summary = employees.copy()
        summary["paid_hours"] = 0.0
        counts = pd.DataFrame(columns=["employee_id", "Nombre de créneaux"])
        service_hours = pd.DataFrame(columns=["employee_id"])
    else:
        corrected = pd.to_numeric(validated["corrected_duration_hours"], errors="coerce")
        duration = pd.to_numeric(validated["duration_hours"], errors="coerce")
        validated.loc[:, "paid_hours"] = corrected.combine_first(duration).fillna(0)
        summary = validated.groupby(["employee_id", "employee"], as_index=False)["paid_hours"].sum()
        counts = validated.groupby("employee_id", as_index=False)["id"].count().rename(columns={"id": "Nombre de créneaux"})
        service_hours = validated.pivot_table(
            index="employee_id",
            columns="service",
            values="paid_hours",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()

    manual_counts = _manual_accepted_counts(manual_requests)
    anomaly_counts = _anomaly_counts(anomalies)
    comments = _manager_comments(validated if not validated.empty else sessions)

    result = employees.merge(summary[["employee_id", "paid_hours"]], on="employee_id", how="left")
    result = result.merge(service_hours, on="employee_id", how="left")
    result = result.merge(counts, on="employee_id", how="left")
    result = result.merge(manual_counts, on="employee_id", how="left")
    result = result.merge(anomaly_counts, on="employee_id", how="left")
    result = result.merge(comments, on="employee_id", how="left")

    for service in ["restaurant", "ménage", "entretien"]:
        if service not in result.columns:
            result[service] = 0.0

    for column in [
        "paid_hours",
        "weekly_target_hours",
        "restaurant",
        "ménage",
        "entretien",
        "Nombre de créneaux",
        "Nombre de demandes manuelles acceptées",
        "Nombre d'anomalies",
    ]:
        result.loc[:, column] = pd.to_numeric(result[column], errors="coerce").fillna(0)
    result.loc[:, "Commentaire responsable"] = result["Commentaire responsable"].fillna("")
    result.loc[:, "Écart"] = result["paid_hours"] - result["weekly_target_hours"]
    result.loc[:, "Statut"] = result["Écart"].map(_hours_status)

    return pd.DataFrame(
        {
            "Employé": result["employee"],
            "Objectif semaine": result["weekly_target_hours"].round(2),
            "Total heures validées": result["paid_hours"].round(2),
            "Écart": result["Écart"].round(2),
            "Statut": result["Statut"],
            "Heures restaurant": result["restaurant"].round(2),
            "Heures ménage": result["ménage"].round(2),
            "Heures entretien": result["entretien"].round(2),
            "Nombre de créneaux": result["Nombre de créneaux"].astype(int),
            "Nombre de demandes manuelles acceptées": result["Nombre de demandes manuelles acceptées"].astype(int),
            "Nombre d'anomalies": result["Nombre d'anomalies"].astype(int),
            "Commentaire responsable": result["Commentaire responsable"],
        },
        columns=SUMMARY_COLUMNS,
    )


def _build_summary_csv(
    sessions: pd.DataFrame,
    manual_requests: pd.DataFrame,
    anomalies: pd.DataFrame,
    employees: pd.DataFrame | None = None,
) -> pd.DataFrame:
    summary = _build_summary(sessions, manual_requests, anomalies, employees)
    if summary.empty:
        return pd.DataFrame(
            columns=[
                *SUMMARY_COLUMNS,
                "objectif_semaine",
                "heures_validees",
                "ecart",
                "statut_heures",
            ]
        )
    result = summary.copy()
    result.loc[:, "objectif_semaine"] = result["Objectif semaine"]
    result.loc[:, "heures_validees"] = result["Total heures validées"]
    result.loc[:, "ecart"] = result["Écart"]
    result.loc[:, "statut_heures"] = result["Statut"]
    return result


def _build_details(sessions: pd.DataFrame) -> pd.DataFrame:
    if sessions.empty:
        return pd.DataFrame(columns=DETAIL_COLUMNS)
    rows = []
    for row in sessions.itertuples(index=False):
        rows.append(
            {
                "Date": _date(row.start_time),
                "Employé": row.employee,
                "Service": row.service,
                "Début": _time(row.start_time),
                "Fin": _time(row.end_time),
                "Durée": _paid_duration(row),
                "Origine": _source_label(row.source, row.validation_status),
                "Statut": _status_label(row.validation_status, row.end_time),
                "Commentaire responsable": row.manager_comment or "",
            }
        )
    return pd.DataFrame(rows, columns=DETAIL_COLUMNS)


def _build_manual_requests(manual_requests: pd.DataFrame) -> pd.DataFrame:
    if manual_requests.empty:
        return pd.DataFrame(columns=MANUAL_COLUMNS)
    rows = []
    for row in manual_requests.itertuples(index=False):
        rows.append(
            {
                "Date": _date(row.requested_start_time),
                "Employé": row.employee,
                "Service": row.service,
                "Début demandé": _time(row.requested_start_time),
                "Fin demandée": _time(row.requested_end_time),
                "Durée demandée": round(float(row.requested_duration_hours or 0), 2),
                "Raison": row.reason or "",
                "Statut": _status_label(row.status, None),
                "Commentaire responsable": row.manager_comment or "",
            }
        )
    return pd.DataFrame(rows, columns=MANUAL_COLUMNS)


def _build_anomalies(anomalies: pd.DataFrame) -> pd.DataFrame:
    if anomalies.empty:
        return pd.DataFrame(columns=ANOMALY_COLUMNS)
    rows = []
    for row in anomalies.itertuples(index=False):
        start = getattr(row, "start_time", "")
        end = getattr(row, "end_time", "")
        duration = getattr(row, "duration_hours", "")
        status = getattr(row, "validation_status", getattr(row, "status", ""))
        rows.append(
            {
                "Date": _date(start),
                "Employé": getattr(row, "employee", ""),
                "Service": getattr(row, "service", ""),
                "Type anomalie": getattr(row, "anomaly", ""),
                "Détail": _anomaly_detail(start, end, duration),
                "Statut": _status_label(status, end),
            }
        )
    return pd.DataFrame(rows, columns=ANOMALY_COLUMNS)


def _build_logs(logs: pd.DataFrame) -> pd.DataFrame:
    if logs.empty:
        return pd.DataFrame(columns=LOG_COLUMNS)
    return pd.DataFrame(
        {
            "Date action": logs["timestamp"].map(_datetime_label),
            "Employé": logs["employee"].fillna(""),
            "Service": logs["service"].fillna(""),
            "Action": logs["action"].map(_status_label).fillna(logs["action"]),
            "Ancienne valeur": logs["old_value"].fillna("").map(_status_label),
            "Nouvelle valeur": logs["new_value"].fillna(""),
            "Commentaire responsable": logs["manager_comment"].fillna(""),
        },
        columns=LOG_COLUMNS,
    )


def _build_activity_logs(logs: pd.DataFrame) -> pd.DataFrame:
    if logs.empty:
        return pd.DataFrame(columns=ACTIVITY_LOG_COLUMNS)
    return pd.DataFrame(
        {
            "Date action": logs["created_at"].map(_datetime_label),
            "Responsable": logs["actor"].fillna(""),
            "Établissement": logs["establishment"].fillna(""),
            "Action": logs["action"].map(_audit_action_label).fillna(logs["action"]),
            "Élément": [
                _audit_entity_label(entity_type, entity_id)
                for entity_type, entity_id in zip(logs["entity_type"], logs["entity_id"], strict=False)
            ],
            "Ancienne valeur": logs["old_value"].fillna(""),
            "Nouvelle valeur": logs["new_value"].fillna(""),
            "Commentaire": logs["comment"].fillna(""),
        },
        columns=ACTIVITY_LOG_COLUMNS,
    )


def _manual_accepted_counts(manual_requests: pd.DataFrame) -> pd.DataFrame:
    if manual_requests.empty:
        return pd.DataFrame(columns=["employee_id", "Nombre de demandes manuelles acceptées"])
    accepted = manual_requests[manual_requests["status"].isin([VALIDATED, CORRECTED])]
    if accepted.empty:
        return pd.DataFrame(columns=["employee_id", "Nombre de demandes manuelles acceptées"])
    return accepted.groupby("employee_id", as_index=False)["id"].count().rename(
        columns={"id": "Nombre de demandes manuelles acceptées"}
    )


def _anomaly_counts(anomalies: pd.DataFrame) -> pd.DataFrame:
    if anomalies.empty or "employee_id" not in anomalies.columns:
        return pd.DataFrame(columns=["employee_id", "Nombre d'anomalies"])
    return anomalies.groupby("employee_id", as_index=False)["anomaly"].count().rename(
        columns={"anomaly": "Nombre d'anomalies"}
    )


def _manager_comments(sessions: pd.DataFrame) -> pd.DataFrame:
    if sessions.empty or "manager_comment" not in sessions.columns:
        return pd.DataFrame(columns=["employee_id", "Commentaire responsable"])
    comments = sessions.dropna(subset=["manager_comment"]).copy()
    if comments.empty:
        return pd.DataFrame(columns=["employee_id", "Commentaire responsable"])
    comments = comments.groupby("employee_id")["manager_comment"].apply(
        lambda values: " | ".join(str(value) for value in values if str(value).strip())
    )
    return comments.reset_index().rename(columns={"manager_comment": "Commentaire responsable"})


def _payable_sessions(sessions: pd.DataFrame) -> pd.DataFrame:
    if sessions.empty:
        return sessions.copy()
    payable = sessions[sessions["validation_status"].isin([VALIDATED, CORRECTED])].copy()
    if payable.empty:
        return payable
    corrected = pd.to_numeric(payable["corrected_duration_hours"], errors="coerce")
    duration = pd.to_numeric(payable["duration_hours"], errors="coerce")
    payable.loc[:, "paid_hours"] = corrected.combine_first(duration).fillna(0)
    return payable


def _employees_for_export(sessions: pd.DataFrame, employees: pd.DataFrame | None) -> pd.DataFrame:
    if employees is not None and not employees.empty:
        result = employees.copy()
    elif sessions.empty:
        result = pd.DataFrame(columns=["employee_id", "employee", "weekly_target_hours"])
    else:
        result = sessions[["employee_id", "employee"]].drop_duplicates().copy()
    if "weekly_target_hours" not in result.columns:
        result.loc[:, "weekly_target_hours"] = 35.0
    result.loc[:, "weekly_target_hours"] = pd.to_numeric(result["weekly_target_hours"], errors="coerce").fillna(35)
    return result.sort_values("employee")


def _number_value(value, default: float = 0) -> float:
    if value is None or pd.isna(value):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _hours_status(diff: float) -> str:
    if diff < -2:
        return "Pas assez d’heures"
    if diff > 2:
        return "Trop d’heures"
    return "Conforme"


def _week_day_by_date(week_start: str) -> dict[str, str]:
    start = pd.to_datetime(week_start).date()
    return {(start + timedelta(days=index)).isoformat(): day for index, day in enumerate(WEEK_DAYS)}


def _session_detail(row) -> str:
    return f"{_time(row.start_time)}-{_time(row.end_time)} {row.service} ({_hours_label(_paid_duration(row))})"


def _hours_label(value: float) -> str:
    hours = float(value or 0)
    if hours.is_integer():
        return f"{int(hours)}h"
    return f"{hours:.2f}h"


def _format_workbook(workbook) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="EAF0F2")
    weekend_fill = PatternFill(fill_type="solid", fgColor="FFF4E5")
    total_fill = PatternFill(fill_type="solid", fgColor="E8F3EA")
    numeric_headers = {
        *WEEK_DAYS,
        "Total semaine",
        "Objectif semaine",
        "Heures validées",
        "Écart",
        "Total heures validées",
        "Heures restaurant",
        "Heures ménage",
        "Heures entretien",
        "Durée",
        "Durée demandée",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "B2"
        if worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            if cell.value in ("Samedi", "Dimanche"):
                cell.fill = weekend_fill
            if cell.value and "Total" in str(cell.value):
                cell.fill = total_fill
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            header = column_cells[0].value
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, max((len(line) for line in value.splitlines()), default=0))
                if "\n" in value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                if header in ("Samedi", "Dimanche") and cell.row > 1:
                    cell.fill = weekend_fill
                if header and "Total" in str(header):
                    cell.font = Font(bold=True)
                if header in numeric_headers and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
                if header == "Statut" and cell.row > 1:
                    if cell.value == "Conforme":
                        cell.fill = PatternFill(fill_type="solid", fgColor="E5F4EC")
                    elif cell.value == "Pas assez d’heures":
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFF3D1")
                    elif cell.value == "Trop d’heures":
                        cell.fill = PatternFill(fill_type="solid", fgColor="F8E6E6")
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"


def _paid_duration(row) -> float:
    corrected = getattr(row, "corrected_duration_hours", None)
    duration = getattr(row, "duration_hours", None)
    value = corrected if pd.notna(corrected) else duration
    return round(float(value or 0), 2)


def _source_label(source: str, status: str) -> str:
    if status == CORRECTED:
        return "Correction responsable"
    if source == "manual":
        return "Demande manuelle"
    return "Scan QR"


def _status_label(status: str, end_time=None) -> str:
    if status == "En attente de validation" and pd.isna(end_time):
        return "Incomplet"
    return {
        "En attente de validation": "En attente",
        "Demande manuelle en attente": "Demande manuelle en attente",
        VALIDATED: "Validé",
        CORRECTED: "Corrigé",
        REFUSED: "Refusé",
        "validate": "Validé",
        "correct": "Corrigé",
        "refuse": "Refusé",
    }.get(status, status or "")


def _audit_action_label(action: str | None) -> str:
    return {
        "work_session_validated": "Créneau validé",
        "work_session_corrected": "Créneau corrigé",
        "work_session_refused": "Créneau refusé",
        "manual_request_accepted": "Demande manuelle acceptée",
        "manual_request_corrected": "Demande manuelle corrigée",
        "manual_request_refused": "Demande manuelle refusée",
        "manager_work_session_created": "Créneau manuel créé",
        "employee_weekly_target_updated": "Objectif hebdomadaire modifié",
        "employee_deactivated": "Personne désactivée",
        "manager_user_created": "Responsable créé",
        "manager_user_deactivated": "Responsable désactivé",
        "export_excel_generated": "Export Excel généré",
        "export_csv_generated": "Export CSV généré",
        "backup_created": "Sauvegarde créée",
        "qr_code_regenerated": "QR code régénéré",
    }.get(str(action or ""), action or "")


def _audit_entity_label(entity_type: str | None, entity_id) -> str:
    entity = {
        "work_session": "Créneau",
        "manual_time_request": "Demande manuelle",
        "employee": "Personne",
        "user": "Responsable",
        "export": "Export",
        "backup": "Sauvegarde",
        "qr_code": "QR code",
    }.get(str(entity_type or ""), entity_type or "Élément")
    if entity_id is None or pd.isna(entity_id):
        return entity
    try:
        return f"{entity} #{int(entity_id)}"
    except (TypeError, ValueError):
        return entity


def _date(value) -> str:
    text = "" if value is None or pd.isna(value) else str(value)
    return text[:10]


def _time(value) -> str:
    if value is None or pd.isna(value) or value == "":
        return ""
    return str(value)[11:16]


def _datetime_label(value) -> str:
    if value is None or pd.isna(value) or value == "":
        return ""
    text = str(value)
    return f"{text[:10]} {text[11:16]}"


def _anomaly_detail(start, end, duration) -> str:
    parts = []
    if start:
        parts.append(f"Début {_time(start)}")
    if end and not pd.isna(end):
        parts.append(f"Fin {_time(end)}")
    if duration is not None and not pd.isna(duration):
        parts.append(f"Durée {round(float(duration), 2)} h")
    return " - ".join(parts)
